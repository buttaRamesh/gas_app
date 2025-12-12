"""
Excel Exporter - generates Excel (.xlsx) export files with formatting.
"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .base import BaseExporter


class ExcelExporter(BaseExporter):
    """
    Export data to Excel format with professional formatting.

    Features:
    - Bold headers with background color
    - Auto-sized columns
    - Borders and alignment
    - Freeze header row
    """

    extension = 'xlsx'
    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def generate(self):
        """
        Generate Excel file and return as HttpResponse.

        Returns:
            HttpResponse with Excel content
        """
        # Create workbook and worksheet
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Export Data"

        # Add title and subtitle
        row_idx = 1

        # Title row
        worksheet.merge_cells(f'A{row_idx}:Z{row_idx}')  # Merge first row
        title_cell = worksheet.cell(row=row_idx, column=1, value=self.page_title)
        title_cell.font = Font(bold=True, size=14, color="003366")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        row_idx += 1

        # Subtitle with total records
        total_records = self.queryset.count()
        worksheet.merge_cells(f'A{row_idx}:Z{row_idx}')
        subtitle_cell = worksheet.cell(row=row_idx, column=1, value=f"Total Records: {total_records}")
        subtitle_cell.font = Font(size=10, color="666666")
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
        row_idx += 1

        # Empty row for spacing
        row_idx += 1

        # Get headers
        headers = self.get_headers()

        # Write headers - values only first
        header_row = row_idx
        for col_idx, header in enumerate(headers, start=1):
            worksheet.cell(row=header_row, column=col_idx, value=header)

        # Track column max widths for auto-sizing
        col_widths = [len(str(header)) for header in headers]

        # Write data rows - VALUES ONLY (no formatting during write)
        row_idx = header_row + 1
        if self.serializer_class:
            # Toggle between ORM and Raw SQL implementations
            USE_RAW_SQL = True  # ← Set to False to use ORM version

            import time
            t_bulk_start = time.time()

            if USE_RAW_SQL:
                from core.exports.bulk_loaders_raw_sql import bulk_load_consumer_export_data_raw_sql
                print(f"  📊 Excel: Using RAW SQL bulk loading...")
                data = bulk_load_consumer_export_data_raw_sql(self.queryset, self.visible_fields)
            else:
                from core.exports.bulk_loaders import bulk_load_consumer_export_data
                print(f"  📊 Excel: Using ORM bulk loading...")
                data = bulk_load_consumer_export_data(self.queryset, self.visible_fields)

            print(f"  ⏱️  Excel: Bulk load complete: {(time.time() - t_bulk_start):.2f}s")

            t_write_start = time.time()
            for row_dict in data:
                for col_idx, field in enumerate(self.visible_fields, start=1):
                    value = self._format_value(row_dict.get(field, ''))
                    worksheet.cell(row=row_idx, column=col_idx, value=value)
                    # Track max width for auto-sizing
                    col_widths[col_idx - 1] = max(col_widths[col_idx - 1], len(str(value)))
                row_idx += 1
            print(f"  ⏱️  Excel: Write {row_idx - 2} rows (values only): {(time.time() - t_write_start):.2f}s")
        else:
            # Fast path: Stream values() directly
            for row_dict in self.queryset.values(*self.visible_fields).iterator(chunk_size=1000):
                for col_idx, field in enumerate(self.visible_fields, start=1):
                    value = self._format_value(row_dict.get(field, ''))
                    worksheet.cell(row=row_idx, column=col_idx, value=value)
                    col_widths[col_idx - 1] = max(col_widths[col_idx - 1], len(str(value)))
                row_idx += 1

        # Apply all formatting in one pass (much faster than per-cell)
        import time
        t_format_start = time.time()

        # Header row styling
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx in range(1, len(headers) + 1):
            cell = worksheet.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        print(f"  ⏱️  Excel: Apply formatting: {(time.time() - t_format_start):.2f}s")

        # Auto-size columns based on calculated widths
        t_resize_start = time.time()
        for col_idx in range(1, len(headers) + 1):
            column_letter = worksheet.cell(row=header_row, column=col_idx).column_letter
            # Set column width (with a reasonable max)
            adjusted_width = min(col_widths[col_idx - 1] + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        print(f"  ⏱️  Excel: Auto-size columns: {(time.time() - t_resize_start):.2f}s")

        # Freeze header row (freeze after title and subtitle)
        worksheet.freeze_panes = f'A{header_row + 1}'

        # Save to BytesIO buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        # Create and return response
        return self.create_response(buffer.read())

    def _format_value(self, value):
        """Format a value for Excel output."""
        if value is None:
            return ''
        if isinstance(value, bool):
            return 'Yes' if value else 'No'
        return value
