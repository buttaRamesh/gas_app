# File: consumers/views.py
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q
from django.http import HttpResponse
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

from core.pagination import CustomPageNumberPagination
from authentication.permissions import HasResourcePermission
from .models import Consumer
from .serializers import (
    ConsumerListSerializer,
    ConsumerDetailSerializer,
    ConsumerCreateUpdateSerializer,
    ConsumersByRouteSerializer,
    ConsumerNewActivationSerializer
)


class ConsumerViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Consumer operations.

    Provides:
    - list: Get all consumers with pagination and filtering
    - retrieve: Get single consumer with full details
    - create: Create new consumer
    - update: Update consumer (PUT)
    - partial_update: Partial update consumer (PATCH)
    - destroy: Delete consumer

    Custom actions:
    - kyc_pending: Get consumers with pending KYC
    - by_route: Get consumers by route code
    - search: Advanced search
    """

    queryset = Consumer.objects.select_related(
        'person',
        'category',
        'consumer_type',
        'bpl_type',
        'dct_type',
        'scheme'
    ).prefetch_related(
        'person__addresses',
        'person__contacts'
    ).all()

    permission_classes = [IsAuthenticated, HasResourcePermission]
    resource_name = 'consumers'
    pagination_class = CustomPageNumberPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'consumer_type', 'opting_status', 'status', 'is_kyc_done', 'scheme']
    search_fields = ['consumer_number', 'person__person_name', 'identification__ration_card_num', 'lpg_id']
    ordering_fields = ['consumer_number', 'person__person_name', 'id']
    ordering = ['consumer_number']
    
    def get_permissions(self):
        """
        Override to use different permission checks for specific actions.
        - by_route action requires 'routes' VIEW permission instead of 'consumers'
        """
        if self.action == 'by_route':
            # For by_route action, create a custom permission class that checks routes permission
            from authentication.permissions import HasPermission

            class HasRoutesViewPermission(HasPermission):
                def has_permission(self, request, view):
                    if not request.user or not request.user.is_authenticated:
                        return False
                    return request.user.has_permission('routes', 'view')

            return [IsAuthenticated(), HasRoutesViewPermission()]

        # Default permissions for all other actions
        return super().get_permissions()

    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'list':
            return ConsumerListSerializer
        elif self.action == 'retrieve':
            return ConsumerDetailSerializer
        else:  # create, update, partial_update
            return ConsumerCreateUpdateSerializer
    
    def get_queryset(self):
        """
        Optionally restricts the returned consumers,
        by filtering against query parameters in the URL.
        """
        queryset = super().get_queryset()
        
        # Example: Filter by KYC status from query params
        kyc_status = self.request.query_params.get('kyc_status', None)
        if kyc_status is not None:
            if kyc_status.lower() == 'true':
                queryset = queryset.filter(is_kyc_done=True)
            elif kyc_status.lower() == 'false':
                queryset = queryset.filter(is_kyc_done=False)
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def kyc_pending(self, request):
        """
        Get all consumers with pending KYC.

        GET /api/consumers/kyc_pending/
        """
        queryset = self.get_queryset().filter(is_kyc_done=False)

        # Apply pagination
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = ConsumerListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = ConsumerListSerializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def new_customers(self, request):
        """
        Get all NEW consumers with completion status info.
        Shows consumer number, name, info (COMPLETED/INCOMPLETE), and mobile.

        GET /api/consumers/new_customers/
        """
        queryset = self.get_queryset().filter(status='NEW')

        # Apply pagination
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = ConsumerNewActivationSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = ConsumerNewActivationSerializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def activate(self, request, pk=None):
        """
        Activate a NEW consumer (only if all required details are complete).

        POST /api/consumers/{id}/activate/
        """
        consumer = self.get_object()

        # Check if consumer is NEW
        if consumer.status != 'NEW':
            return Response(
                {'error': f'Cannot activate consumer with status {consumer.get_status_display()}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check if all required details are complete
        missing = []
        if not (consumer.person and consumer.person.addresses.exists()):
            missing.append('address')
        if not (consumer.person and consumer.person.contacts.exists()):
            missing.append('contact')

        # Check identification
        from .models import Identification
        try:
            if not consumer.identification:
                missing.append('identification')
        except Identification.DoesNotExist:
            missing.append('identification')

        # Check category and consumer_type
        if not consumer.category or not consumer.consumer_type:
            missing.append('additional details')

        if missing:
            return Response(
                {
                    'error': 'Cannot activate incomplete consumer',
                    'missing': missing,
                    'message': f'Missing: {", ".join(missing)}'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # Activate the consumer
        consumer.status = 'ACTIVE'
        consumer.save()

        serializer = ConsumerDetailSerializer(consumer)
        return Response({
            'message': 'Consumer activated successfully',
            'consumer': serializer.data
        })
    
    @action(detail=False, methods=['get'])
    def by_route(self, request):
        """
        Get consumers by route code or route ID with detailed information.

        GET /api/consumers/by_route/?route_code=R001
        GET /api/consumers/by_route/?route_id=1
        GET /api/consumers/by_route/?route_id=1&search=john

        Returns paginated list with consumer details including:
        - consumer_id, consumer_number, consumer_name
        - mobile, address, route_code
        - category, consumer_type, cylinders count

        Supports search parameter to filter by consumer_number or consumer_name

        Requires: VIEW permission on 'routes' resource
        """
        route_code = request.query_params.get('route_code', None)
        route_id = request.query_params.get('route_id', None)
        search = request.query_params.get('search', None)

        if not route_code and not route_id:
            return Response(
                {'error': 'Either route_code or route_id parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get queryset with optimized queries to avoid N+1 issues
        queryset = Consumer.objects.select_related(
            'person',
            'category',
            'consumer_type',
            'route_assignment__route'
        ).prefetch_related(
            'person__contacts',
            'person__addresses',
            'connections'
        )

        if route_id:
            queryset = queryset.filter(route_assignment__route_id=route_id)
        else:
            queryset = queryset.filter(route_assignment__route__area_code=route_code)

        # Apply search filter if provided
        if search:
            queryset = queryset.filter(
                Q(consumer_number__icontains=search) |
                Q(person__person_name__icontains=search)
            )

        # Always use pagination for this endpoint
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = ConsumersByRouteSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = ConsumersByRouteSerializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def route(self, request, pk=None):
        """
        Get route assignment for a specific consumer.

        GET /api/consumers/{id}/route/
        """
        consumer = self.get_object()

        try:
            assignment = consumer.route_assignment
            return Response({
                'route_id': assignment.route.id,
                'route_code': assignment.route.area_code,
                'route_description': assignment.route.area_code_description
            })
        except:
            return Response(
                {'message': 'No route assigned to this consumer'},
                status=status.HTTP_404_NOT_FOUND
            )

    @action(detail=True, methods=['post'])
    def assign_route(self, request, pk=None):
        """
        Assign a route to a consumer.

        POST /api/consumers/{id}/assign_route/
        Body: { "route_id": 1 }
        """
        consumer = self.get_object()

        # Check if consumer has a consumer_number
        if not consumer.consumer_number:
            return Response(
                {'error': 'Consumer number is required before assigning a route'},
                status=status.HTTP_400_BAD_REQUEST
            )

        route_id = request.data.get('route_id')

        if not route_id:
            return Response(
                {'error': 'route_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from routes.models import Route
            route = Route.objects.get(id=route_id)

            # Create or update assignment
            from consumers.models import ConsumerRouteAssignment
            assignment, created = ConsumerRouteAssignment.objects.update_or_create(
                consumer=consumer,
                defaults={'route': route}
            )

            return Response({
                'message': 'Route assigned successfully',
                'route_id': route.id,
                'route_code': route.area_code,
                'route_description': route.area_code_description
            })
        except Route.DoesNotExist:
            return Response(
                {'error': 'Route not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['patch'])
    def update_route(self, request, pk=None):
        """
        Update/change route assignment for a consumer.

        PATCH /api/consumers/{id}/update_route/
        Body: { "route_id": 2 }
        """
        consumer = self.get_object()

        # Check if consumer has a consumer_number
        if not consumer.consumer_number:
            return Response(
                {'error': 'Consumer number is required before updating route'},
                status=status.HTTP_400_BAD_REQUEST
            )

        route_id = request.data.get('route_id')

        if not route_id:
            return Response(
                {'error': 'route_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from routes.models import Route
            from consumers.models import ConsumerRouteAssignment
            route = Route.objects.get(id=route_id)

            # Update existing assignment
            assignment = consumer.route_assignment
            assignment.route = route
            assignment.save()

            return Response({
                'message': 'Route updated successfully',
                'route_id': route.id,
                'route_code': route.area_code,
                'route_description': route.area_code_description
            })
        except Route.DoesNotExist:
            return Response(
                {'error': 'Route not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except ConsumerRouteAssignment.DoesNotExist:
            return Response(
                {'error': 'No route assignment found for this consumer'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['patch'])
    def update_kyc_status(self, request, pk=None):
        """
        Update only the KYC status of a consumer.
        
        PATCH /api/consumers/{id}/update_kyc_status/
        Body: { "is_kyc_done": true }
        """
        consumer = self.get_object()
        is_kyc_done = request.data.get('is_kyc_done')
        
        if is_kyc_done is None:
            return Response(
                {'error': 'is_kyc_done field is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        consumer.is_kyc_done = is_kyc_done
        consumer.save()
        
        serializer = ConsumerDetailSerializer(consumer)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """
        Get consumer statistics.
        
        GET /api/consumers/statistics/
        """
        total = self.get_queryset().count()
        kyc_done = self.get_queryset().filter(is_kyc_done=True).count()
        kyc_pending = total - kyc_done
        
        by_status = {}
        for status_choice in Consumer.OptingStatus.choices:
            count = self.get_queryset().filter(opting_status=status_choice[0]).count()
            by_status[status_choice[1]] = count
        
        return Response({
            'total_consumers': total,
            'kyc_done': kyc_done,
            'kyc_pending': kyc_pending,
            'by_opting_status': by_status,
        })

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """
        Export consumers to CSV file.

        GET /api/consumers/export_csv/?search=john&ordering=consumer_number&is_kyc_done=true
        """
        # Get filtered queryset
        queryset = self.filter_queryset(self.get_queryset())

        # Create CSV response
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="consumers_export.csv"'

        # Add BOM for Excel UTF-8 compatibility
        response.write('\ufeff')

        writer = csv.writer(response)

        # Write header
        writer.writerow([
            'Consumer Number',
            'Consumer Name',
            'Category',
            'Type',
            'Opting Status',
            'KYC Done',
            'Mobile',
            'Ration Card',
            'Aadhar',
            'PAN',
            'LPG ID',
        ])

        # Write data
        for consumer in queryset:
            contact = consumer.person.contacts.first() if consumer.person else None
            try:
                identification = consumer.identification
                ration_card = identification.ration_card_num or ''
                aadhar = identification.aadhar_num or ''
                pan = identification.pan_num or ''
            except:
                ration_card = ''
                aadhar = ''
                pan = ''

            writer.writerow([
                consumer.consumer_number,
                consumer.person.person_name if consumer.person else 'Unknown',
                consumer.category.name if consumer.category else '',
                consumer.consumer_type.name if consumer.consumer_type else '',
                consumer.get_opting_status_display(),
                'Yes' if consumer.is_kyc_done else 'No',
                contact.mobile_number if contact else '',
                ration_card,
                aadhar,
                pan,
                consumer.lpg_id or '',
            ])

        return response

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """
        Export consumers to Excel file.

        GET /api/consumers/export_excel/?search=john&ordering=consumer_number&is_kyc_done=true
        """
        # Get filtered queryset
        queryset = self.filter_queryset(self.get_queryset())

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Consumers"

        # Define header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write header
        headers = [
            'Consumer Number',
            'Consumer Name',
            'Category',
            'Type',
            'Opting Status',
            'KYC Done',
            'Mobile',
            'Ration Card',
            'Aadhar',
            'PAN',
            'LPG ID',
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write data
        for row_num, consumer in enumerate(queryset, 2):
            contact = consumer.person.contacts.first() if consumer.person else None
            try:
                identification = consumer.identification
                ration_card = identification.ration_card_num or ''
                aadhar = identification.aadhar_num or ''
                pan = identification.pan_num or ''
            except:
                ration_card = ''
                aadhar = ''
                pan = ''

            ws.cell(row=row_num, column=1, value=consumer.consumer_number)
            ws.cell(row=row_num, column=2, value=consumer.person.person_name if consumer.person else 'Unknown')
            ws.cell(row=row_num, column=3, value=consumer.category.name if consumer.category else '')
            ws.cell(row=row_num, column=4, value=consumer.consumer_type.name if consumer.consumer_type else '')
            ws.cell(row=row_num, column=5, value=consumer.get_opting_status_display())
            ws.cell(row=row_num, column=6, value='Yes' if consumer.is_kyc_done else 'No')
            ws.cell(row=row_num, column=7, value=contact.mobile_number if contact else '')
            ws.cell(row=row_num, column=8, value=ration_card)
            ws.cell(row=row_num, column=9, value=aadhar)
            ws.cell(row=row_num, column=10, value=pan)
            ws.cell(row=row_num, column=11, value=consumer.lpg_id or '')

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="consumers_export.xlsx"'
        wb.save(response)

        return response

    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        """
        Export consumers to PDF file.

        GET /api/consumers/export_pdf/?search=john&ordering=consumer_number&is_kyc_done=true
        """
        # Get filtered queryset
        queryset = self.filter_queryset(self.get_queryset())

        # Create PDF response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="consumers_export.pdf"'

        # Create PDF document
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []

        # Add title
        styles = getSampleStyleSheet()
        title = Paragraph("<b>Consumers Export</b>", styles['Title'])
        elements.append(title)

        # Prepare table data
        data = [[
            'Consumer #',
            'Name',
            'Category',
            'Type',
            'Status',
            'KYC',
            'Mobile',
        ]]

        for consumer in queryset:
            contact = consumer.person.contacts.first() if consumer.person else None
            consumer_name = consumer.person.person_name if consumer.person else 'Unknown'
            data.append([
                consumer.consumer_number,
                consumer_name[:30],  # Truncate long names
                consumer.category.name if consumer.category else '',
                consumer.consumer_type.name if consumer.consumer_type else '',
                consumer.get_opting_status_display()[:10],  # Truncate
                'Yes' if consumer.is_kyc_done else 'No',
                contact.mobile_number if contact else '',
            ])

        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)

        # Get PDF from buffer
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)

        return response